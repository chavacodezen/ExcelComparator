import tkinter as tk
from tkinter import filedialog, messagebox
import os
import time
from openpyxl import load_workbook


class ExcelComparatorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Comparador de Archivos Excel")

        # Define the main background color variable
        self.bg_color = "#d0d0d0"

        # Set window size and padding
        self.root.geometry("720x480")
        self.root.config(padx=30, pady=30)

        # Initialize file paths and default columns
        self.main_file = ''
        self.compare_file = ''
        self.main_col = 'B'  # Default main column to compare
        self.compare_col = 'C'  # Default compare column
        self.insert_col = 'G'  # Default column for inserting data
        self.data_col = 'N'  # Default data column in compare file

        # Create UI components
        self.create_widgets()

    def create_widgets(self):
        # Create a frame for the file selection section
        file_selection_frame = tk.Frame(
            self.root, bg=self.bg_color, padx=15, pady=5)
        file_selection_frame.grid(row=0, column=0, sticky="ew", pady=(5, 5))

        # File selection: Main file
        tk.Label(file_selection_frame, text="Selecciona el archivo donde se insertarán los resultados",
                 bg=self.bg_color).grid(row=0, column=0, sticky="w", pady=(10, 5))
        self.main_file_button = tk.Button(
            file_selection_frame, text="Seleccionar archivo", command=self.select_main_file)
        self.main_file_button.grid(
            row=0, column=1, sticky="w", pady=(10, 5), padx=(10, 0))

        self.main_file_label = tk.Label(
            file_selection_frame, text="Selecciona el archivo de destino", width=30, anchor="w", bg=self.bg_color)
        self.main_file_label.grid(row=1, column=1, sticky="w", pady=(0, 10))

        # File selection: Compare file
        tk.Label(file_selection_frame, text="Selecciona el archivo a comparar",
                 bg=self.bg_color).grid(row=2, column=0, sticky="w", pady=(10, 5))
        self.compare_file_button = tk.Button(
            file_selection_frame, text="Seleccionar archivo", command=self.select_compare_file)
        self.compare_file_button.grid(
            row=2, column=1, sticky="w", pady=(10, 5), padx=(10, 0))
        self.compare_file_label = tk.Label(
            file_selection_frame, text="Selecciona el archivo fuente", width=30, anchor="w", bg=self.bg_color)
        self.compare_file_label.grid(row=3, column=1, sticky="w", pady=(0, 10))

        # Column selection frame (Reduced right-side space further)
        column_selection_frame = tk.Frame(
            self.root, bg=self.bg_color, padx=15, pady=5)
        column_selection_frame.grid(row=1, column=0, sticky="ew", pady=(10, 5))

        tk.Label(column_selection_frame, text="Columna principal (archivo destino):",
                 bg=self.bg_color).grid(row=0, column=0, sticky="w", pady=(5, 3))
        self.main_col_entry = self.create_entry(
            column_selection_frame, self.main_col)
        self.main_col_entry.grid(
            row=0, column=1, sticky="w", pady=(5, 3), padx=(10, 0))

        tk.Label(column_selection_frame, text="Columna de comparación (archivo fuente):",
                 bg=self.bg_color).grid(row=1, column=0, sticky="w", pady=(5, 3))
        self.compare_col_entry = self.create_entry(
            column_selection_frame, self.compare_col)
        self.compare_col_entry.grid(
            row=1, column=1, sticky="w", pady=(5, 3), padx=(10, 0))

        # Column insert frame (Reduced right-side space further)
        column_insert_frame = tk.Frame(
            self.root, bg=self.bg_color, padx=15, pady=5)
        column_insert_frame.grid(row=2, column=0, sticky="ew", pady=(10, 5))

        tk.Label(column_insert_frame, text="Columna para insertar datos:",
                 bg=self.bg_color).grid(row=0, column=0, sticky="w", pady=(5, 3))
        self.insert_col_entry = self.create_entry(
            column_insert_frame, self.insert_col)
        self.insert_col_entry.grid(
            row=0, column=1, sticky="w", pady=(5, 3), padx=(10, 0))

        tk.Label(column_insert_frame, text="Columna con los datos (archivo fuente):",
                 bg=self.bg_color).grid(row=1, column=0, sticky="w", pady=(5, 3))
        self.data_col_entry = self.create_entry(
            column_insert_frame, self.data_col)
        self.data_col_entry.grid(
            row=1, column=1, sticky="w", pady=(5, 3), padx=(10, 0))

        # Start process button
        self.start_button = tk.Button(
            self.root, text="Iniciar proceso", command=self.start_process)
        self.start_button.grid(row=3, column=0, columnspan=2, pady=(20, 10))

    def create_entry(self, parent_frame, default_value):
        """
        Create an entry field that only allows letters and converts input to uppercase.
        """
        entry = tk.Entry(parent_frame)
        entry.insert(0, default_value)

        # Binding the input to allow only letters and uppercase them
        entry.bind("<KeyRelease>", lambda event,
                   entry=entry: self.validate_and_uppercase(entry, event))
        return entry

    def validate_and_uppercase(self, entry, event):
        """
        Validate that only letters are entered and convert input to uppercase.
        """
        input_text = entry.get()

        # Allow only letters (uppercase/lowercase)
        if not input_text.isalpha() and input_text != "":
            # Remove invalid character
            entry.delete(len(input_text) - 1, tk.END)
            return

        # Convert text to uppercase
        entry.delete(0, tk.END)
        entry.insert(0, input_text.upper())

    def select_main_file(self):
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel Files", "*.xlsx")])
        if file_path:
            self.main_file = file_path
            # Update label with selected file name
            self.main_file_label.config(text=os.path.basename(file_path))

    def select_compare_file(self):
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel Files", "*.xlsx")])
        if file_path:
            self.compare_file = file_path
            self.compare_file_label.config(text=os.path.basename(
                file_path))  # Update label with selected file name

    def start_process(self):
        if not self.main_file or not self.compare_file:
            messagebox.showerror(
                "Error", "Por favor, selecciona ambos archivos.")
            return

        try:
            # Start processing
            start_time = time.time()
            self.compare_and_insert_data()
            elapsed_time = time.time() - start_time

            messagebox.showinfo(
                "Proceso completado", f"El proceso finalizó correctamente en {elapsed_time:.2f} segundos.")
        except Exception as e:
            messagebox.showerror("Error", f"Ocurrió un error: {e}")

    def compare_and_insert_data(self):
        main_wb = load_workbook(self.main_file)
        compare_wb = load_workbook(self.compare_file)

        main_sheet = main_wb.active
        compare_sheet = compare_wb.active

        main_col_idx = ord(self.main_col_entry.get().upper()) - ord('A') + 1
        compare_col_idx = ord(
            self.compare_col_entry.get().upper()) - ord('A') + 1
        insert_col_idx = ord(
            self.insert_col_entry.get().upper()) - ord('A') + 1
        data_col_idx = ord(self.data_col_entry.get().upper()) - ord('A') + 1

        for main_row in range(2, main_sheet.max_row + 1):  # Assuming headers in row 1
            main_value = main_sheet.cell(
                row=main_row, column=main_col_idx).value
            inserted_data = None

            # Assuming headers in row 1
            for compare_row in range(2, compare_sheet.max_row + 1):
                compare_value = compare_sheet.cell(
                    row=compare_row, column=compare_col_idx).value
                if main_value == compare_value:
                    inserted_data = compare_sheet.cell(
                        row=compare_row, column=data_col_idx).value
                    break

            main_sheet.cell(row=main_row, column=insert_col_idx,
                            value=inserted_data)

        main_wb.save(self.main_file)
        main_wb.close()
        compare_wb.close()


# Create the main window
root = tk.Tk()
app = ExcelComparatorApp(root)
root.mainloop()
